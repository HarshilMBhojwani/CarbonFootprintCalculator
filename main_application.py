'''
Final Project
Restaurant Carbon_emission Calcualtor
by HArshil Bhojwani
24/04/2023

'''
import openpyxl #to read iin the excel file
import matplotlib.pyplot as plt#matplot is imported to draw charts
from paper import Paper
from raw_meat import RawFood
from plastic import Plastic
from waste_food import Waste
from gas import Gas
#the required classes are printed above
def main():
    print("Welcome to No food Carbon")
    print("We will help you calculate your carbon foot print")
    print("Lets start with you giving us a few basic details about your restaurant company")
    name=input("Enter the name of your Organisation: \n")
    city=input("Enter your city of operation: \n")
    cusine=input("Enter the cusine which is served in your restaurant: \n")
    covers=input("Enter the total number of covers in your restaurant: \n")
    table_turnover=input("Enter the table turover for your restaurant: \n")
    print("Please upload the Excel file we provided you with all your details")
    file_name=input("Enter the file name\n")#this input will be the file name from which data has to nimported
    ct=openpyxl.load_workbook(file_name)#the excel file is called out using openpyxel 
    sh1=ct["Sheet1"]
    #print(ct.active.title)
    print("We will read the data from the sheet now "\
          "and help you caculate your carbon footprint")
    #beef2=(sh1['B11'].value)
    beef=(sh1.cell(11,2).value)#being a standard format for excel all the cells are fixed, only the values are suppose to change 
    #print(beef2)
    seafood=float(sh1.cell(12,2).value)#the value for everything particular data is imported
    poultry=float(sh1.cell(13,2).value)#for the few lines all the data raw produce data is importe
    pork=float(sh1.cell(14,2).value)
    lamb=float(sh1.cell(15,2).value)
    eggs=float(sh1.cell(16,2).value)
    dairy=float(sh1.cell(17,2).value)
    veggie=float(sh1.cell(18,2).value)
    cheese=float(sh1.cell(19,2).value)
    choco=float(sh1.cell(20,2).value)
    coffee=float(sh1.cell(21,2).value)
    cook_oil=float(sh1.cell(22,2).value)
    raw_carbon=RawFood(beef, seafood, poultry, pork, lamb, eggs, dairy, veggie, cheese, choco, coffee, cook_oil )
    #the imported raw produe data is been given to the class, to calculate the all the raw produce carbon emission.
    #Below until line 65 all the methods to calculate the carbon_emission of every raw produce are called. 
    beef_carbon=raw_carbon.redmeat()
    #print(beef_carbon)
    sea_carbon=raw_carbon.seafood_carbon()
    poul_carbon=raw_carbon.poultry_carbon()
    pork_carbon=raw_carbon.pork_carbon()
    lamb_carbon=raw_carbon.carbon_lamb()
    egg_carbon=raw_carbon.carbon_eggs()
    dairy_carbon=raw_carbon.dairy_carbon()
    veggie_carbon=raw_carbon.vegetable_carbon()
    cheese_carbon=raw_carbon.cheese_carbon()
    choco_carbon=raw_carbon.chocolate_carbon()
    coffee_carbon=raw_carbon.coffee_carbon()
    cook_oil=raw_carbon.oil_carbon()
    
    #raw_dict=raw_carbon.raw_total(beef_carbon, sea_carbon,pork_carbon, lamb_carbon, egg_carbon, dairy_carbon, veggie_carbon, cheese_carbon, choco_carbon, coffee_carbon, cook_oil)
    #print(raw_dict)
    #total_raw=sea_carbon+beef_carbon+poul_carbon+pork_carbon+lamb_carbon+egg_carbon+dairy_carbon+veggie_carbon+cheese_carbon+choco_carbon+coffee_carbon+cook_oil
    #all the carbon emission data is been added to a list to plot the chart
    list_raw=[sea_carbon, beef_carbon, poul_carbon, pork_carbon, lamb_carbon, egg_carbon, dairy_carbon, veggie_carbon, cheese_carbon, choco_carbon, coffee_carbon,cook_oil]
    raw_names=["Seafood", "Beef", "Poultry", "Pork", "Lamb", "Eggs", "Dairy", "Vegetables",
                "Cheese", "Chocolate", "Coffee", "Cooking oil"]
    #all list of names for every raw produce to ass labels to the chart
    
    plt.barh(raw_names,list_raw)#a horizonatal bar chart is to be plotted is called out here
    plt.title("Greenhouse Gas emission from Raw Produce")#the title for the chart is given
    plt.ylabel("Raw Produce Variety")#y axis label is given
    plt.xlabel("Greenhouse emission of each type in C02 kilogram")#x axis label is given
    plt.show()#the chart is printed

    foodwaste=float(sh1.cell(25,2).value)#similar to above method, foodwaste data is called imported
    waste_class=Waste(foodwaste)#to calucalte the carbon emissions its been attached to class waste
    waste_carbon=waste_class.carbon_foodwaste()#method is calucalted to give the required output
    sum_raw=sum(list_raw)#the raw food carbon_emission is ummed to compare the data eith food wastage
    waste_list=[waste_carbon, sum_raw]#to plot a chart they are beeinn given to the list
    waste_label=["Food Wastage", "Raw Produce"]#the label list for the chart
    a=["Green", "Red"]

    plt.barh(waste_label, waste_list, color=a)#to plot the bar chart
    plt.title("Greenhouse Gas emission from Foodwastage vs Total Fresh Raw Produce")#title for the graph

    plt.xlabel("Greenhouse emission in Co2 kilogram")#the required label for the graph
    plt.show()
    #to calculate the data of plastic different data as per the excel is called out
    pet=float(sh1.cell(35,2).value)
    ldpe=float(sh1.cell(40,2).value)
    pvc=float(sh1.cell(44,2).value)
    ps=float(sh1.cell(52,2).value)
    paper_roll=float(sh1.cell(56,2).value)
    tissue=float(sh1.cell(55,2).value)
    paper_bag=float(sh1.cell(54,2).value)
    recycle_plastic=float(sh1.cell(59,2).value)
    recycle_paper=float(sh1.cell(60,2).value)
    foil=float(sh1.cell(62,2).value)
    #plastic_carbon=Plastic(pet, pvc, hdpe, ps, recycle_plastic, foil )
    plastic_carbon=Plastic(pet, pvc, ldpe, ps, recycle_plastic, foil)#to  calucalte the carbon emission of each variety the plastic class is called out.
    #using different class methodscarbon emission for plastic type is calculated
    carbon_pet=(plastic_carbon.carbon_pet())
    carbon_pvc=(plastic_carbon.carbon_pvc())
    carbon_ldpe=(plastic_carbon.carbon_ldpe())
    carbon_ps=(plastic_carbon.carbon_ps())
    recycle_carbon=(plastic_carbon.plastic_recycled())
    foil_carbon=(plastic_carbon.carbon_foil())
    #Data obatined is transfered to al list to plot graphs
    list_plastic=[carbon_pet, carbon_pvc, carbon_ldpe, carbon_ps, recycle_carbon, foil_carbon]
    plastic_names=["PET", "PVC", "LDPE", "PS", "Recycle", "Foil"]#list labels are put into a list again for graph purpoe
    c="cyan"#color for the graph
    plt.barh(plastic_names, list_plastic, color=c)#graph is plotted
    plt.title("Greenhouse Gas emission from Plastic products and its derivitives and savings by recyling")#a header for the graph is givem
    plt.xlabel("Greenhouse emission in Co2 kilogram")#the x axis and yaxis laves are provided
    plt.ylabel("Plastic variety")
    plt.show()

    paper_carbon=Paper(paper_roll, tissue, paper_bag, recycle_paper)#a class is called out to calculate the carbon emission of paper products
    roll_carbon=paper_carbon.carbon_roll()
    tissue_carbon=paper_carbon.carbon_tissue()
    bag_paper_carbon=paper_carbon.paper_bag_carbon()
    carbon_paper_recycle=paper_carbon.recycle_paper_carbon()
#mehtods are called to calculate emission of different paper

# a list is created using the to plot the data 
    paper_list=[roll_carbon, tissue_carbon, bag_paper_carbon, carbon_paper_recycle]#rto plot a graph all the data is tranferred to a list
    paper_names=["Carbon Emission Paperroll", "Carbon Emission Tissue", "Carbon Emission Paperbag", "Emission saved from recycling"]#label for the graph is created
    b="orange"
    plt.barh(paper_names, paper_list, color=b)#the graph is plotted
    plt.title("Greenhouse Gas emission from Paper products and savings by recyling")
    plt.xlabel("Greenhouse emission in Co2 kg")#title for the graph, and x, y axis label sif given
    plt.ylabel("Paper variety")
    plt.show()
#lastly eletricity for gas and electricity data is extracted from the excel
    electricity=float(sh1.cell(29,2).value)
    #print(electricity)
    electricity_form=(sh1.cell(4,5).value)
    #print(electricity_form)
    cooking_gas=float(sh1.cell(28,2).value)
    wood=float(sh1.cell(65,2).value)
    coal=float(sh1.cell(66,2).value)
    #class is called of calucalte carbon emission of  d
    energy_class=Gas(cooking_gas, electricity, coal, wood)
    if electricity_form =="coal" or electricity_form == "Coal" or electricity_form == "COAL":
        electricity_carbon=energy_class.carbon_coal()#condition statments are given to check fo the source of the electricity
    elif electricity_form == "gas" or electricity_form== "Gas" or electricity_form == "GAS":
        electricity_carbon=energy_class.carbon_natural_gas()
    elif electricity_form == "oil" or electricity_form== "Oil" or electricity_form == "OIL":
        electricity_carbon=energy_class.carbon_oil()
#methods are called out to calculate the carbon emission each category
    cooking_carbon=energy_class.carbon_natural_gas()
    carbon_wood=energy_class.wood_carbon()
    carbon_coal=energy_class.alternative_coal()
    energy_list=[electricity_carbon, cooking_carbon, carbon_wood, carbon_coal]#the data is transferred to a list to plot a graph
    energy_names=["Electricity", "Cooking_gas", "Wood", "Coal" ]#labels fro the graph is created
    d="purple"#colour for the graph is given
    plt.barh(energy_names, energy_list, color=d)#graph is plotted
    plt.title("Greenhouse Gas emission from different Energy Sources")
    plt.xlabel("Greenhouse emission of each type of Energy Source in Kg")#title and axis labels are given
    plt.ylabel("Energy Forms")
    plt.show()

    sum_plastic=(sum(list_plastic))-recycle_plastic#each sector carbon emisison  is summed to compare the carbon emsiion of each category
    sum_paper=(sum(paper_list))-recycle_paper
    sum_energy=sum(energy_list)

    total_carbon_list=[sum_raw, foodwaste, sum_plastic, sum_paper, sum_energy]#the summed catgory data is transfered to a list to plot a final graphs comparing every data
    total_names=["Raw Produce", "Foodwaste", "Plastic", "Paper", "Energy"]#the labels for the graphs is given
    f=["red", "green", "yellow", "blue", "black"]#color for every bar is different for better raeasilbility
    plt.barh(total_names, total_carbon_list, color=f)#the graph is plotted
    plt.title("Greehouse emission contribution of each core section")
    plt.xlabel("Greenhouse emisson Co2 per kg")#labels and titles are given
    plt.ylabel("Sectors")
    plt.show()




if __name__ == "__main__":

 main()

